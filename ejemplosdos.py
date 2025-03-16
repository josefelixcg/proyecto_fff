import tkinter as tkr
from tkinter import messagebox
import openpyxl as opp
import pandas as pd

baseexcel=pd.ExcelFile("BASE_MATERIALES_APPSHEETS.xlsx")
def btn_hojas():
    mesam = opp.load_workbook("BASE_MATERIALES_APPSHEETS.xlsx")
    qq=mesam.sheetnames
    messagebox.showinfo("HOJAS",message=f"El libro tiene las hojas \n {qq}")
    
base=tkr.Tk()
base.title("DESDE EXCEL")

ff=tkr.Button(base,text="Hojas del Libro",command=btn_hojas)
ff.pack()

#baseexcel=opp.load_workbook("BASE_MATERIALES_APPSHEETS.xlsx")

#hojas=baseexcel.sheetnames

#usar_label = tkr.Label(base, text=f"hojas del libro de excel a trabajar \n {hojas}",font=("Arial",20))
#usar_label.pack(pady=30)

base.mainloop()

#DATOS = pd.read_excel(r"https://docs.google.com/spreadsheets/d/1-9xC-kPfziG4o2ZU-eS3YAnMYKZd31lm/edit?usp=sharing&ouid=110144339162798547112&rtpof=true&sd=true","Inspecciones")
#print(DATOS.head(5))

df = pd.read_excel('BASE_MATERIALES_APPSHEETS.xlsx',sheet_name="Inspecciones")
# Seleccionar la columna 'A' y convertirla en una lista
columna_a = df["A"].tolist()

# Mostrar la lista
print(columna_a)
